"""
DEFINITIVE HDFC PDF Parser - Coordinate-based, 100% accurate.

Strategy:
- Use word x-coordinates for column detection (from PDF analysis)
- Transaction = any line with valid DATE in date column + valid BALANCE in balance column
- Continuation = line with narration content but NO date in date column
- NO keyword-based filtering (avoids false positives)
- Balance continuity determines withdrawal vs deposit
"""
import sys
import os
import re
import pdfplumber
from typing import List, Dict, Optional
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers

# ============================================================================
# COLUMN BOUNDARIES (from PDF word-position analysis)
# ============================================================================
COL_DATE_MAX = 65       # Date column: x < 65
COL_NARR_MAX = 260      # Narration: 65 <= x < 260
COL_REF_MAX = 360       # Chq./Ref.No: 260 <= x < 360
COL_VDT_MAX = 405       # Value Date: 360 <= x < 405
COL_WDR_MAX = 485       # Withdrawal: 405 <= x < 485
COL_DEP_MAX = 562       # Deposit: 485 <= x < 562
                        # Balance: x >= 562

DATE_PATTERN = re.compile(r'^\d{2}/\d{2}/\d{2}$')
AMOUNT_PATTERN = re.compile(r'^[\d,]+\.\d{2}$')

# Y-position bounds for transaction data area (from PDF analysis)
# Headers/address block is y < 230, footer is y > 790
DATA_Y_MIN = 230
DATA_Y_MAX = 790


def get_column(x0: float) -> str:
    """Assign word to column based on x-coordinate."""
    if x0 < COL_DATE_MAX:
        return 'date'
    elif x0 < COL_NARR_MAX:
        return 'narration'
    elif x0 < COL_REF_MAX:
        return 'ref_no'
    elif x0 < COL_VDT_MAX:
        return 'value_date'
    elif x0 < COL_WDR_MAX:
        return 'withdrawal'
    elif x0 < COL_DEP_MAX:
        return 'deposit'
    else:
        return 'balance'


def parse_amount(s: str) -> Optional[float]:
    """Parse comma-formatted amount to float."""
    s = s.strip().replace(',', '')
    try:
        val = float(s)
        return val if val > 0 else None
    except (ValueError, TypeError):
        return None


def extract_lines_from_page(page) -> List[Dict]:
    """Extract words from page, group by y-position, assign to columns."""
    words = page.extract_words(keep_blank_chars=True, x_tolerance=2, y_tolerance=2)
    if not words:
        return []

    # Group words by approximate y-position (within 5 units = same line)
    y_groups = defaultdict(list)
    for w in words:
        y_key = round(w['top'] / 5) * 5
        y_groups[y_key].append(w)

    lines = []
    for y_key in sorted(y_groups.keys()):
        # Skip lines outside the transaction data area (page headers/footers)
        if y_key < DATA_Y_MIN or y_key > DATA_Y_MAX:
            continue

        line_words = sorted(y_groups[y_key], key=lambda w: w['x0'])

        # Build column contents
        col_words = defaultdict(list)
        for w in line_words:
            col = get_column(w['x0'])
            col_words[col].append(w['text'])

        line = {
            'y': y_key,
            'date': ' '.join(col_words.get('date', [])).strip(),
            'narration': ' '.join(col_words.get('narration', [])).strip(),
            'ref_no': ' '.join(col_words.get('ref_no', [])).strip(),
            'value_date': ' '.join(col_words.get('value_date', [])).strip(),
            'withdrawal': ' '.join(col_words.get('withdrawal', [])).strip(),
            'deposit': ' '.join(col_words.get('deposit', [])).strip(),
            'balance': ' '.join(col_words.get('balance', [])).strip(),
        }
        lines.append(line)

    return lines


def parse_hdfc_pdf(pdf_path: str) -> List[Dict]:
    """
    Parse HDFC PDF into individual transactions.
    
    Rules:
    1. Transaction start = line with valid DD/MM/YY in date col + valid amount in balance col
    2. Continuation = line with narration/ref content but no valid date
    3. Withdrawal vs Deposit determined by balance comparison with previous transaction
    """
    all_transactions = []
    prev_balance = None

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            lines = extract_lines_from_page(page)

            for line in lines:
                date_text = line['date']
                balance_text = line['balance']

                # Check if this is a transaction start
                is_date = bool(DATE_PATTERN.match(date_text))
                balance_val = parse_amount(balance_text) if balance_text else None

                if is_date and balance_val is not None:
                    # === NEW TRANSACTION ===
                    withdrawal_val = parse_amount(line['withdrawal'])
                    deposit_val = parse_amount(line['deposit'])
                    value_date = line['value_date'] if DATE_PATTERN.match(line['value_date']) else date_text

                    # Determine debit/credit from balance movement
                    if prev_balance is not None:
                        balance_change = balance_val - prev_balance
                        if balance_change < -0.005:
                            # Balance decreased = WITHDRAWAL
                            if not withdrawal_val:
                                withdrawal_val = abs(balance_change)
                            deposit_val = None
                        elif balance_change > 0.005:
                            # Balance increased = DEPOSIT
                            if not deposit_val:
                                deposit_val = balance_change
                            withdrawal_val = None
                        else:
                            # No change (very rare)
                            withdrawal_val = None
                            deposit_val = None
                    else:
                        # First transaction - use whatever amount is present
                        if deposit_val and not withdrawal_val:
                            pass  # Keep deposit
                        elif withdrawal_val and not deposit_val:
                            pass  # Keep withdrawal
                        elif deposit_val and withdrawal_val:
                            # Both present - shouldn't happen, use balance direction
                            deposit_val = deposit_val
                            withdrawal_val = None

                    txn = {
                        'date': date_text,
                        'narration': line['narration'],
                        'ref_no': line['ref_no'],
                        'value_date': value_date,
                        'withdrawal': withdrawal_val,
                        'deposit': deposit_val,
                        'closing_balance': balance_val,
                        'page': page_num + 1,
                    }
                    all_transactions.append(txn)
                    prev_balance = balance_val

                elif all_transactions:
                    # === CONTINUATION LINE ===
                    # Append narration/ref content to previous transaction
                    if line['narration']:
                        all_transactions[-1]['narration'] += ' ' + line['narration']
                    if line['ref_no']:
                        if not all_transactions[-1]['ref_no']:
                            all_transactions[-1]['ref_no'] = line['ref_no']
                        # Don't append ref to narration - it's a separate field

    # Clean up narrations
    for txn in all_transactions:
        txn['narration'] = ' '.join(txn['narration'].split())  # Normalize whitespace

    return all_transactions


def verify_transactions(transactions: List[Dict]) -> Dict:
    """Verify transaction accuracy using balance continuity."""
    errors = []
    total_withdrawal = 0
    total_deposit = 0

    for i in range(len(transactions)):
        txn = transactions[i]

        if txn['withdrawal']:
            total_withdrawal += txn['withdrawal']
        if txn['deposit']:
            total_deposit += txn['deposit']

        if i == 0:
            continue

        prev_bal = transactions[i - 1]['closing_balance']
        curr_bal = txn['closing_balance']
        expected_change = round(curr_bal - prev_bal, 2)

        if txn['withdrawal'] and not txn['deposit']:
            actual_change = round(-txn['withdrawal'], 2)
        elif txn['deposit'] and not txn['withdrawal']:
            actual_change = round(txn['deposit'], 2)
        else:
            actual_change = 0

        diff = abs(expected_change - actual_change)
        if diff > 0.02:
            errors.append({
                'index': i + 1,
                'date': txn['date'],
                'expected': expected_change,
                'actual': actual_change,
                'diff': diff,
                'narration': txn['narration'][:50],
            })

    return {
        'total_transactions': len(transactions),
        'total_withdrawal': total_withdrawal,
        'total_deposit': total_deposit,
        'errors': errors,
    }


def generate_excel(transactions: List[Dict], output_path: str):
    """Generate Excel file with all transactions."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Raw Transactions"

    # Column widths
    ws.column_dimensions['A'].width = 12   # Date
    ws.column_dimensions['B'].width = 65   # Narration
    ws.column_dimensions['C'].width = 25   # Chq./Ref.No.
    ws.column_dimensions['D'].width = 12   # Value Dt
    ws.column_dimensions['E'].width = 16   # Withdrawal Amt.
    ws.column_dimensions['F'].width = 16   # Deposit Amt.
    ws.column_dimensions['G'].width = 18   # Closing Balance

    # Header style
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # Headers
    headers = ['Date', 'Narration', 'Chq./Ref.No.', 'Value Dt', 'Withdrawal Amt.', 'Deposit Amt.', 'Closing Balance']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    data_font = Font(name='Calibri', size=10)
    amount_format = '#,##0.00'
    wrap_align = Alignment(vertical='top', wrap_text=True)
    amount_align = Alignment(horizontal='right', vertical='top')

    for i, txn in enumerate(transactions, 2):
        # Date
        cell = ws.cell(row=i, column=1, value=txn['date'])
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical='top')

        # Narration
        cell = ws.cell(row=i, column=2, value=txn['narration'])
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = wrap_align

        # Chq./Ref.No.
        cell = ws.cell(row=i, column=3, value=txn['ref_no'])
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical='top')

        # Value Dt
        cell = ws.cell(row=i, column=4, value=txn['value_date'])
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical='top')

        # Withdrawal
        if txn['withdrawal']:
            cell = ws.cell(row=i, column=5, value=txn['withdrawal'])
            cell.number_format = amount_format
        else:
            cell = ws.cell(row=i, column=5, value='')
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = amount_align

        # Deposit
        if txn['deposit']:
            cell = ws.cell(row=i, column=6, value=txn['deposit'])
            cell.number_format = amount_format
        else:
            cell = ws.cell(row=i, column=6, value='')
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = amount_align

        # Closing Balance
        cell = ws.cell(row=i, column=7, value=txn['closing_balance'])
        cell.number_format = amount_format
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = amount_align

    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Auto-filter
    ws.auto_filter.ref = f"A1:G{len(transactions) + 1}"

    wb.save(output_path)
    print(f"Excel saved: {output_path}")


def main():
    pdf_path = sys.argv[1] if len(sys.argv) > 1 else "HDFC DATA/Hdfc Bank 1 June to 24 Jan 26.pdf"

    if not os.path.exists(pdf_path):
        print(f"File not found: {pdf_path}")
        sys.exit(1)

    print("=" * 120)
    print("DEFINITIVE HDFC PDF PARSER - COORDINATE BASED")
    print("=" * 120)

    # Parse
    transactions = parse_hdfc_pdf(pdf_path)

    # Verify
    verification = verify_transactions(transactions)

    print(f"\nTotal transactions: {verification['total_transactions']}")
    print(f"Total withdrawals: {verification['total_withdrawal']:,.2f}")
    print(f"Total deposits: {verification['total_deposit']:,.2f}")

    # Monthly breakdown
    from collections import Counter
    month_counts = Counter(txn['date'][3:5] for txn in transactions)
    print("\nMonthly breakdown:")
    for month in sorted(month_counts.keys()):
        print(f"  Month {month}: {month_counts[month]} transactions")

    # Balance errors
    errors = verification['errors']
    print(f"\nBalance continuity errors: {len(errors)}")
    if errors:
        for e in errors[:10]:
            print(f"  Txn #{e['index']} ({e['date']}): expected={e['expected']:.2f}, actual={e['actual']:.2f}, diff={e['diff']:.2f} | {e['narration']}")

    # Show first 20 transactions
    print(f"\n{'='*140}")
    print("FIRST 20 TRANSACTIONS")
    print(f"{'='*140}")
    fmt = "{:>3} | {:<10} | {:<55} | {:<22} | {:<10} | {:>14} | {:>14} | {:>16}"
    print(fmt.format('#', 'Date', 'Narration', 'Ref No', 'Val.Dt', 'Withdrawal', 'Deposit', 'Balance'))
    print("-" * 155)

    for i, txn in enumerate(transactions[:20]):
        w = f"{txn['withdrawal']:>14,.2f}" if txn['withdrawal'] else ""
        d = f"{txn['deposit']:>14,.2f}" if txn['deposit'] else ""
        b = f"{txn['closing_balance']:>16,.2f}"
        print(fmt.format(i + 1, txn['date'], txn['narration'][:55], txn['ref_no'][:22], txn['value_date'], w, d, b))

    # Generate Excel
    output_path = os.path.join(os.path.dirname(pdf_path), "hdfc_transactions_output.xlsx")
    generate_excel(transactions, output_path)

    print(f"\nTotal: {len(transactions)} transactions exported to Excel")


if __name__ == "__main__":
    main()

"""
Airco Insights — Upload Router (Bank-Specific Architecture)
===============================================================
POST /process endpoint for accuracy-first processing.

Key Features:
- Bank selection is REQUIRED (no auto-detection)
- Routes to bank-specific processors
- Returns detailed validation metrics
- 100% validation before response

Frontend Flow:
1. User enters: Full Name, Account Type, Bank Name
2. User uploads PDF
3. User selects mode (Free/Hybrid)
4. System processes using bank-specific logic
5. Returns validated, categorized results
"""

import logging
import os
from typing import Optional

from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse
import pikepdf

from app.core.security import validate_upload_file, validate_file_size
from app.utils.file_handler import save_temp_file, get_temp_dir
from app.services.pipeline_orchestrator import (
    process_statement as run_pipeline,
    PipelineValidationError,
    PipelineAbortError,
    UnsupportedBankError,
)

logger = logging.getLogger(__name__)

router = APIRouter()

# Sheet keys aligned to actual Excel sheet order (8 sheets)
SHEET_KEYS = [
    "account_summary",    # Sheet 1 — Summary
    "monthly_analysis",   # Sheet 2 — Monthly Analysis
    "weekly_analysis",    # Sheet 3 — Weekly Analysis
    "category_analysis",  # Sheet 4 — Category Analysis
    "bounces_penal",      # Sheet 5 — Bounces & Penal
    "funds_received",     # Sheet 6 — Funds Received
    "funds_remittance",   # Sheet 7 — Funds Remittance
    "raw_transactions",   # Sheet 8 — Raw Transaction
]

MAX_PREVIEW_ROWS = 50


def _extract_sheet_previews(excel_path: str) -> dict:
    """Extract sheet preview data for frontend."""
    previews = {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        
        for idx, sheet_name in enumerate(wb.sheetnames):
            if idx >= len(SHEET_KEYS):
                break
            
            ws = wb[sheet_name]
            rows_data = []
            headers = []
            
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                str_row = [str(cell) if cell is not None else "" for cell in row]
                if row_idx == 0:
                    headers = str_row
                else:
                    if row_idx <= MAX_PREVIEW_ROWS:
                        rows_data.append(str_row)
            
            previews[SHEET_KEYS[idx]] = {
                "title": f"Sheet {idx + 1} — {sheet_name}",
                "headers": headers,
                "rows": rows_data,
            }
        
        wb.close()
    except Exception as e:
        logger.warning("Could not extract sheet previews: %s", str(e))
    
    return previews


def _build_frontend_response(result: dict, mode: str) -> dict:
    """Transform result to frontend-compatible format."""
    excel_path = result.get("excel_path", "")
    
    # Convert file path to download URL
    excel_filename = os.path.basename(excel_path) if excel_path else ""
    excel_url = f"/download/{excel_filename}" if excel_filename else ""
    
    # Extract sheet previews
    previews = {}
    if excel_path and os.path.isfile(excel_path):
        previews = _extract_sheet_previews(excel_path)
    
    response = {
        "status": result.get("status", "success"),
        "mode": mode,
        "bank": result.get("bank_key", ""),
        "excel_url": excel_url,
        "stats": result.get("stats", {}),
        "validation": result.get("validation", {}),
        "performance": result.get("performance", {}),
        **previews,
    }
    
    # Add error info if present
    if result.get("error"):
        response["error"] = result["error"]
    
    return response


@router.post("/process")
async def process_statement_endpoint(
    file: UploadFile = File(...),
    full_name: str = Form(""),
    account_type: str = Form("Salaried"),
    bank_name: str = Form(...),  # Required!
    mode: str = Form("free"),
    api_key: Optional[str] = Form(None),
    pdf_password: Optional[str] = Form(None),  # For password-protected PDFs
):
    """
    Process bank statement using accuracy-first architecture.
    
    This endpoint uses bank-specific processors for maximum accuracy.
    Bank name is REQUIRED - no auto-detection.
    
    Args:
        file: PDF bank statement file
        full_name: Account holder name
        account_type: "Salaried" or "Business"
        bank_name: Bank name (required) - e.g., "HDFC Bank"
        mode: "free" (rule engine only) or "hybrid" (rule + AI)
        api_key: Anthropic API key (required for hybrid mode)
    
    Returns:
        JSON with processed data, validation status, and download URL
    """
    temp_pdf_path = None
    
    try:
        # 1. Validate bank name
        if not bank_name or not bank_name.strip():
            raise HTTPException(
                status_code=400,
                detail="Bank name is required. Please select a bank."
            )
        
        # 2. Validate upload
        validate_upload_file(file)
        content = await validate_file_size(file)
        logger.info(
            "Upload received: %s (%d bytes) bank=%s mode=%s",
            file.filename, len(content), bank_name, mode
        )
        
        # 3. Save to temp file
        temp_pdf_path = save_temp_file(content, extension=".pdf")
        
        # 4. Check if PDF is password-protected and unlock if needed
        password_check = _check_pdf_password(temp_pdf_path, pdf_password)
        if password_check["is_locked"]:
            if not pdf_password:
                logger.warning("PDF is password-protected but no password provided")
                raise HTTPException(
                    status_code=400,
                    detail={
                        "error": "This PDF is password-protected. Please provide the password to unlock it.",
                        "stage": "validation",
                        "code": "PASSWORD_PROTECTED",
                        "is_locked": True,
                    }
                )
            elif password_check["decrypted_path"]:
                # Password worked, use decrypted file
                temp_pdf_path = password_check["decrypted_path"]
                logger.info("PDF unlocked successfully, using decrypted file")
            else:
                # Password failed
                logger.warning("Incorrect password provided for locked PDF")
                raise HTTPException(
                    status_code=400,
                    detail={
                        "error": password_check["error"] or "Incorrect password provided",
                        "stage": "validation",
                        "code": "INCORRECT_PASSWORD",
                    }
                )
        
        # 5. Build user_info
        user_info = {
            "full_name": full_name,
            "account_type": account_type,
            "bank_name": bank_name,
        }
        
        # 5. Get output directory
        output_dir = get_temp_dir()
        
        # 6. Run processing pipeline
        result = run_pipeline(
            file_path=temp_pdf_path,
            user_info=user_info,
            mode=mode,
            api_key=api_key,
            output_dir=output_dir,
        )

        if result.get("status") != "success":
            err = result.get("error") or {}
            error_message = (
                err.get("message")
                if isinstance(err, dict)
                else None
            ) or "Could not process this PDF. Please verify the statement format and try again."

            raise HTTPException(
                status_code=400,
                detail={
                    "error": error_message,
                    "stage": err.get("stage", "processing") if isinstance(err, dict) else "processing",
                    "code": err.get("code", "PROCESSING_FAILED") if isinstance(err, dict) else "PROCESSING_FAILED",
                }
            )

        stats = result.get("stats") or {}
        if int(stats.get("total_transactions", 0) or 0) <= 0:
            bank_name = user_info.get("bank_name", "bank").lower()
            raise HTTPException(
                status_code=400,
                detail={
                    "error": f"Could not extract any transactions from this {bank_name.title()} statement. Please upload a text-based PDF downloaded from internet banking.",
                    "stage": "parsing",
                    "code": "NO_TRANSACTIONS",
                }
            )
        
        # 7. Build frontend response
        return _build_frontend_response(result, mode)
    
    except PipelineValidationError as e:
        logger.warning("Validation error: %s", str(e))
        raise HTTPException(
            status_code=400,
            detail={
                "error": str(e),
                "stage": e.stage,
                "code": e.error_code,
            }
        )
    
    except UnsupportedBankError as e:
        logger.warning("Unsupported bank: %s", str(e))
        raise HTTPException(
            status_code=400,
            detail={
                "error": str(e),
                "stage": e.stage,
                "code": e.error_code,
                "supported_banks": ["HDFC Bank"],  # Update as more banks are added
            }
        )
    
    except PipelineAbortError as e:
        logger.error("Pipeline abort: %s", str(e))
        raise HTTPException(
            status_code=400,
            detail={
                "error": str(e),
                "stage": e.stage,
                "code": e.error_code,
            }
        )
    
    except HTTPException:
        raise
    
    except Exception as e:
        logger.error("Unexpected error: %s", str(e), exc_info=True)
        raise HTTPException(
            status_code=500,
            detail={
                "error": "An unexpected error occurred while processing your statement.",
                "stage": "unknown",
                "code": "UNEXPECTED_ERROR",
            }
        )


def _check_pdf_password(file_path: str, password: str = None) -> dict:
    """
    Check if PDF is password-protected and try to decrypt.
    
    Returns:
        {"is_locked": bool, "decrypted_path": str or None, "error": str or None}
    """
    try:
        # Try to open without password
        pdf = pikepdf.open(file_path)
        pdf.close()
        return {"is_locked": False, "decrypted_path": file_path, "error": None}
    except pikepdf.PasswordError:
        # PDF is password-protected
        if not password:
            return {"is_locked": True, "decrypted_path": None, "error": "PDF is password-protected"}
        
        # Try with provided password
        try:
            pdf = pikepdf.open(file_path, password=password)
            # Save decrypted version
            decrypted_path = file_path.replace(".pdf", "_decrypted.pdf")
            pdf.save(decrypted_path)
            pdf.close()
            return {"is_locked": True, "decrypted_path": decrypted_path, "error": None}
        except pikepdf.PasswordError:
            return {"is_locked": True, "decrypted_path": None, "error": "Incorrect password"}
        except Exception as e:
            return {"is_locked": True, "decrypted_path": None, "error": f"Decryption failed: {str(e)}"}
    except Exception as e:
        return {"is_locked": False, "decrypted_path": file_path, "error": None}


@router.post("/check-pdf")
async def check_pdf_status(
    file: UploadFile = File(...),
):
    """
    Check if uploaded PDF is password-protected.
    
    Returns:
        {"is_locked": bool, "filename": str}
    """
    try:
        validate_upload_file(file)
        content = await validate_file_size(file)
        
        # Save to temp file
        temp_pdf_path = save_temp_file(content, extension=".pdf")
        
        # Check if locked
        result = _check_pdf_password(temp_pdf_path)
        
        return {
            "is_locked": result["is_locked"],
            "filename": file.filename,
            "temp_path": temp_pdf_path,
        }
    except Exception as e:
        logger.error("Error checking PDF: %s", str(e))
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/unlock-pdf")
async def unlock_pdf(
    temp_path: str = Form(...),
    password: str = Form(...),
):
    """
    Attempt to unlock a password-protected PDF.
    
    Returns:
        {"success": bool, "error": str or None, "decrypted_path": str or None}
    """
    try:
        if not os.path.exists(temp_path):
            raise HTTPException(status_code=400, detail="File not found")
        
        result = _check_pdf_password(temp_path, password)
        
        if result["decrypted_path"]:
            return {
                "success": True,
                "error": None,
                "decrypted_path": result["decrypted_path"],
            }
        else:
            return {
                "success": False,
                "error": result["error"] or "Failed to unlock PDF",
                "decrypted_path": None,
            }
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error unlocking PDF: %s", str(e))
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/supported-banks")
async def get_supported_banks():
    """
    Get list of supported banks.
    
    Returns:
        List of supported banks with their status
    """
    return {
        "banks": [
            {
                "key": "hdfc",
                "name": "HDFC Bank",
                "status": "available",
                "accuracy": "99%+",
            },
            {
                "key": "axis",
                "name": "Axis Bank",
                "status": "available",
                "accuracy": "99%+",
            },
            {
                "key": "icici",
                "name": "ICICI Bank",
                "status": "available",
                "accuracy": "99%+",
            },
            {
                "key": "kotak",
                "name": "Kotak Mahindra Bank",
                "status": "available",
                "accuracy": "99%+",
            },
            {
                "key": "sbi",
                "name": "SBI",
                "status": "coming_soon",
                "accuracy": None,
            },
        ],
        "default_mode": "free",
        "modes": [
            {
                "key": "free",
                "name": "Free Mode",
                "description": "Deterministic rule engine only. No AI, no API cost.",
            },
            {
                "key": "hybrid",
                "name": "Hybrid Mode",
                "description": "Rule engine + Claude AI for unclassified transactions.",
                "requires_api_key": True,
            },
        ],
    }

"""
Airco Insights — HDFC Excel Generator
=========================================
Generates structured Excel export for HDFC transactions.

6 Sheets:
1. Summary — Monthly breakdown with balances
2. Category Analysis — Credit/Debit categories side-by-side
3. Weekly Analysis — Week-by-week credit/debit split
4. Recurring Analysis — Recurring vs Non-recurring
5. Raw Transactions — Complete transaction list
6. Finbit — Monthly Finbit analysis (30 financial metrics)

Design: Structured export only. No charts, no AI text, no analytics.
"""

import logging
from typing import List, Dict, Any, Optional
from datetime import datetime
from collections import defaultdict

logger = logging.getLogger(__name__)


class HDFCExcelGenerator:
    """Excel report generator for HDFC transactions."""
    
    def __init__(self):
        self.logger = logging.getLogger(f"{__name__}.{self.__class__.__name__}")
    
    def generate(
        self,
        transactions: List[Dict[str, Any]],
        aggregation: Any,
        user_info: Dict[str, Any],
        output_path: str,
    ) -> str:
        """Generate Excel report matching exact specifications."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl library required")
        
        self.logger.info("Generating HDFC Excel: %s", output_path)
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Styles
        HEADER_FONT = Font(bold=True, size=11)
        HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
        
        BORDER = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        CENTER = Alignment(horizontal="center", vertical="center")
        LEFT = Alignment(horizontal="left", vertical="center")
        RIGHT = Alignment(horizontal="right", vertical="center")
        
        CURRENCY = '₹#,##0.00'
        
        # Extract metadata
        account_no = user_info.get("account_no", "N/A")
        full_name = user_info.get("full_name", "N/A")
        
        # Get date range from transactions
        if transactions:
            dates = [t.get("date") for t in transactions if t.get("date")]
            stmt_from = min(dates) if dates else "N/A"
            stmt_to = max(dates) if dates else "N/A"
        else:
            stmt_from = stmt_to = "N/A"
        
        # ========================================
        # SHEET 1: Summary
        # ========================================
        ws1 = wb.create_sheet("Summary")
        
        # Header block
        ws1['A1'] = "Name"
        ws1['B1'] = "Account No"
        ws1['C1'] = "Statement From"
        ws1['D1'] = "Statement To"
        
        for col in ['A', 'B', 'C', 'D']:
            ws1[f'{col}1'].font = HEADER_FONT
            ws1[f'{col}1'].border = BORDER
        
        ws1['A2'] = full_name
        ws1['B2'] = account_no
        ws1['C2'] = stmt_from
        ws1['D2'] = stmt_to
        
        for col in ['A', 'B', 'C', 'D']:
            ws1[f'{col}2'].border = BORDER
        
        # Monthly table
        monthly_data = self._aggregate_monthly(transactions)
        
        if monthly_data:
            # Headers
            row = 4
            ws1.cell(row, 1, "Month").font = HEADER_FONT_WHITE
            ws1.cell(row, 1).fill = HEADER_FILL
            ws1.cell(row, 1).alignment = CENTER
            ws1.cell(row, 1).border = BORDER
            
            col = 2
            for month in monthly_data.keys():
                cell = ws1.cell(row, col, month)
                cell.font = HEADER_FONT_WHITE
                cell.fill = HEADER_FILL
                cell.alignment = CENTER
                cell.border = BORDER
                col += 1
            
            # Rows
            metrics = [
                ("Total credit count", "cr_count"),
                ("Total credit amount", "cr_amt"),
                ("Total debit count", "dr_count"),
                ("Total debit amount", "dr_amt"),
                ("Avg balance", "avg_bal"),
                ("Min balance", "min_bal"),
                ("Max balance", "max_bal"),
                ("Start of month balance", "start_bal"),
                ("End of month balance", "end_bal"),
            ]
            
            row = 5
            for label, key in metrics:
                ws1.cell(row, 1, label).font = HEADER_FONT
                ws1.cell(row, 1).border = BORDER
                
                col = 2
                for month_data in monthly_data.values():
                    val = month_data.get(key, 0)
                    cell = ws1.cell(row, col, val)
                    cell.border = BORDER
                    cell.alignment = CENTER
                    
                    if key in ['cr_amt', 'dr_amt', 'avg_bal', 'min_bal', 'max_bal', 'start_bal', 'end_bal']:
                        cell.number_format = CURRENCY
                    
                    col += 1
                row += 1
        
        # Freeze top row
        ws1.freeze_panes = 'A5'
        
        # Auto width
        for col_idx in range(1, len(monthly_data) + 2):
            ws1.column_dimensions[get_column_letter(col_idx)].width = 18
        
        # ========================================
        # SHEET 2: Category Analysis
        # ========================================
        ws2 = wb.create_sheet("Category Analysis")
        
        # Credit categories (left)
        credit_cats = self._aggregate_credit_categories(transactions)
        
        ws2.cell(1, 1, "Credit Categories").font = HEADER_FONT_WHITE
        ws2.cell(1, 1).fill = HEADER_FILL
        ws2.cell(1, 1).border = BORDER
        ws2.merge_cells('A1:B1')
        
        row = 2
        for cat, amount in credit_cats.items():
            ws2.cell(row, 1, cat).border = BORDER
            cell = ws2.cell(row, 2, amount)
            cell.border = BORDER
            cell.number_format = CURRENCY
            cell.alignment = RIGHT
            row += 1
        
        # Total
        total_cr = sum(credit_cats.values())
        ws2.cell(row, 1, "Total credit amount").font = HEADER_FONT
        ws2.cell(row, 1).border = BORDER
        cell = ws2.cell(row, 2, total_cr)
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.number_format = CURRENCY
        cell.alignment = RIGHT
        
        # Debit categories (right)
        debit_cats = self._aggregate_debit_categories(transactions)
        
        ws2.cell(1, 4, "Debit Categories").font = HEADER_FONT_WHITE
        ws2.cell(1, 4).fill = HEADER_FILL
        ws2.cell(1, 4).border = BORDER
        ws2.merge_cells('D1:E1')
        
        row = 2
        for cat, amount in debit_cats.items():
            ws2.cell(row, 4, cat).border = BORDER
            cell = ws2.cell(row, 5, amount)
            cell.border = BORDER
            cell.number_format = CURRENCY
            cell.alignment = RIGHT
            row += 1
        
        # Total
        total_dr = sum(debit_cats.values())
        ws2.cell(row, 4, "Total debit amount").font = HEADER_FONT
        ws2.cell(row, 4).border = BORDER
        cell = ws2.cell(row, 5, total_dr)
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.number_format = CURRENCY
        cell.alignment = RIGHT
        
        # Column widths
        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 18
        ws2.column_dimensions['D'].width = 25
        ws2.column_dimensions['E'].width = 18
        
        # ========================================
        # SHEET 3: Weekly Analysis
        # ========================================
        ws3 = wb.create_sheet("Weekly Analysis")
        
        # Credit weekly (left)
        weekly_cr = self._aggregate_weekly_credit(transactions)
        
        ws3.cell(1, 1, "Credit weekly split").font = HEADER_FONT_WHITE
        ws3.cell(1, 1).fill = HEADER_FILL
        ws3.cell(1, 1).border = BORDER
        ws3.merge_cells('A1:B1')
        
        row = 2
        for week_label, amount in weekly_cr.items():
            ws3.cell(row, 1, week_label).border = BORDER
            cell = ws3.cell(row, 2, amount)
            cell.border = BORDER
            cell.number_format = CURRENCY
            cell.alignment = RIGHT
            row += 1
        
        # Debit weekly (right)
        weekly_dr = self._aggregate_weekly_debit(transactions)
        
        ws3.cell(1, 4, "Debit weekly split").font = HEADER_FONT_WHITE
        ws3.cell(1, 4).fill = HEADER_FILL
        ws3.cell(1, 4).border = BORDER
        ws3.merge_cells('D1:E1')
        
        row = 2
        for week_label, amount in weekly_dr.items():
            ws3.cell(row, 4, week_label).border = BORDER
            cell = ws3.cell(row, 5, amount)
            cell.border = BORDER
            cell.number_format = CURRENCY
            cell.alignment = RIGHT
            row += 1
        
        # Column widths
        ws3.column_dimensions['A'].width = 20
        ws3.column_dimensions['B'].width = 18
        ws3.column_dimensions['D'].width = 20
        ws3.column_dimensions['E'].width = 18
        
        # ========================================
        # SHEET 4: Recurring Analysis
        # ========================================
        ws4 = wb.create_sheet("Recurring Analysis")
        
        recurring_stats = self._aggregate_recurring(transactions)
        
        # Credit
        ws4.cell(1, 1, "Credit").font = HEADER_FONT_WHITE
        ws4.cell(1, 1).fill = HEADER_FILL
        ws4.cell(1, 1).border = BORDER
        ws4.merge_cells('A1:C1')
        
        ws4.cell(2, 1, "Recurring").border = BORDER
        ws4.cell(2, 2, "Non-recurring").border = BORDER
        ws4.cell(2, 3, "Total").border = BORDER
        
        cell = ws4.cell(3, 1, recurring_stats['cr_recurring'])
        cell.border = BORDER
        cell.number_format = CURRENCY
        cell.alignment = RIGHT
        
        cell = ws4.cell(3, 2, recurring_stats['cr_non_recurring'])
        cell.border = BORDER
        cell.number_format = CURRENCY
        cell.alignment = RIGHT
        
        cell = ws4.cell(3, 3, recurring_stats['cr_total'])
        cell.border = BORDER
        cell.number_format = CURRENCY
        cell.alignment = RIGHT
        
        # Debit
        ws4.cell(5, 1, "Debit").font = HEADER_FONT_WHITE
        ws4.cell(5, 1).fill = HEADER_FILL
        ws4.cell(5, 1).border = BORDER
        ws4.merge_cells('A5:C5')
        
        ws4.cell(6, 1, "Recurring").border = BORDER
        ws4.cell(6, 2, "Non-recurring").border = BORDER
        ws4.cell(6, 3, "Total").border = BORDER
        
        cell = ws4.cell(7, 1, recurring_stats['dr_recurring'])
        cell.border = BORDER
        cell.number_format = CURRENCY
        cell.alignment = RIGHT
        
        cell = ws4.cell(7, 2, recurring_stats['dr_non_recurring'])
        cell.border = BORDER
        cell.number_format = CURRENCY
        cell.alignment = RIGHT
        
        cell = ws4.cell(7, 3, recurring_stats['dr_total'])
        cell.border = BORDER
        cell.number_format = CURRENCY
        cell.alignment = RIGHT
        
        # Column widths
        for col_idx in range(1, 4):
            ws4.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # ========================================
        # SHEET 5: Raw Transactions (HDFC PDF Format)
        # ========================================
        ws5 = wb.create_sheet("Raw Transactions")
        
        # ========================================
        # STATEMENT SUMMARY (extracted from PDF)
        # ========================================
        current_row = 1
        
        # Calculate summary from transactions
        opening_bal = aggregation.opening_balance if aggregation else 0
        closing_bal = aggregation.closing_balance if aggregation else 0
        total_dr_count = sum(1 for t in transactions if (t.get('debit') or 0) > 0)
        total_cr_count = sum(1 for t in transactions if (t.get('credit') or 0) > 0)
        total_debits = sum((t.get('debit') or 0) for t in transactions)
        total_credits = sum((t.get('credit') or 0) for t in transactions)
        
        # Summary header
        ws5.cell(current_row, 1, "STATEMENT SUMMARY").font = Font(bold=True, size=12, color="FFFFFF")
        ws5.cell(current_row, 1).fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        ws5.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        current_row += 1
        
        # Summary column headers
        summary_headers = ["Opening Balance", "Dr Count", "Cr Count", "Debits", "Credits", "Closing Bal"]
        for col_idx, header in enumerate(summary_headers, 1):
            cell = ws5.cell(current_row, col_idx, header)
            cell.font = HEADER_FONT
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = CENTER
            cell.border = BORDER
        current_row += 1
        
        # Summary values
        summary_values = [opening_bal, total_dr_count, total_cr_count, total_debits, total_credits, closing_bal]
        for col_idx, value in enumerate(summary_values, 1):
            cell = ws5.cell(current_row, col_idx, value)
            cell.border = BORDER
            cell.alignment = CENTER
            if col_idx in [1, 4, 5, 6]:  # Amount columns
                cell.number_format = CURRENCY
        current_row += 2  # Add spacing
        
        # ========================================
        # TRANSACTION TABLE
        # ========================================
        # Headers - Exact HDFC format
        headers = ["Date", "Narration", "Chq./Ref.No.", "Value Dt", "Withdrawal Amt.", "Deposit Amt.", "Closing Balance"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws5.cell(current_row, col_idx, header)
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = BORDER
        
        # Data
        row = current_row + 1
        for txn in transactions:
            # Date
            ws5.cell(row, 1, txn.get('date', '')).border = BORDER
            ws5.cell(row, 1).alignment = CENTER
            
            # Narration (full description)
            narration = txn.get('description', '')
            ws5.cell(row, 2, narration).border = BORDER
            ws5.cell(row, 2).alignment = LEFT
            
            # Chq./Ref.No.
            ref_no = txn.get('ref_no', '')
            ws5.cell(row, 3, ref_no).border = BORDER
            ws5.cell(row, 3).alignment = CENTER
            
            # Value Dt
            value_date = txn.get('value_date', '')
            ws5.cell(row, 4, value_date).border = BORDER
            ws5.cell(row, 4).alignment = CENTER
            
            # Withdrawal Amt. (Debit)
            debit_val = txn.get('debit')
            if debit_val:
                cell = ws5.cell(row, 5, debit_val)
                cell.border = BORDER
                cell.number_format = CURRENCY
                cell.alignment = RIGHT
            else:
                ws5.cell(row, 5, '').border = BORDER
            
            # Deposit Amt. (Credit)
            credit_val = txn.get('credit')
            if credit_val:
                cell = ws5.cell(row, 6, credit_val)
                cell.border = BORDER
                cell.number_format = CURRENCY
                cell.alignment = RIGHT
            else:
                ws5.cell(row, 6, '').border = BORDER
            
            # Closing Balance
            cell = ws5.cell(row, 7, txn.get('balance', 0))
            cell.border = BORDER
            cell.number_format = CURRENCY
            cell.alignment = RIGHT
            
            row += 1
        
        # Freeze at transaction header row (after summary section)
        freeze_row = current_row + 1
        ws5.freeze_panes = f'A{freeze_row}'
        
        # Column widths - Match HDFC format
        ws5.column_dimensions['A'].width = 12   # Date
        ws5.column_dimensions['B'].width = 50   # Narration
        ws5.column_dimensions['C'].width = 18   # Chq./Ref.No.
        ws5.column_dimensions['D'].width = 12   # Value Dt
        ws5.column_dimensions['E'].width = 18   # Withdrawal Amt.
        ws5.column_dimensions['F'].width = 18   # Deposit Amt.
        ws5.column_dimensions['G'].width = 18   # Closing Balance
        
        # ========================================
        # SHEET 6: Finbit Analysis
        # ========================================
        ws6 = wb.create_sheet("Finbit")
        
        finbit_months, finbit_data = self._compute_finbit_monthly(
            transactions, opening_bal
        )
        
        if finbit_months and finbit_data:
            # Title row
            ws6.cell(1, 1, "FINBIT ANALYSIS").font = Font(
                bold=True, size=13, color="FFFFFF")
            ws6.cell(1, 1).fill = HEADER_FILL
            ws6.merge_cells(
                start_row=1, start_column=1,
                end_row=1, end_column=len(finbit_months) + 1)
            ws6.cell(1, 1).alignment = CENTER
            
            # Account info
            ws6.cell(2, 1, (
                f"Account: {account_no}  |  Name: {full_name}"
                f"  |  Period: {stmt_from} to {stmt_to}"
            ))
            ws6.cell(2, 1).font = Font(italic=True, size=10)
            ws6.merge_cells(
                start_row=2, start_column=1,
                end_row=2, end_column=len(finbit_months) + 1)
            
            # Header row (row 4)
            hdr_row = 4
            hdr_c = ws6.cell(hdr_row, 1, "Metric")
            hdr_c.font = HEADER_FONT_WHITE
            hdr_c.fill = HEADER_FILL
            hdr_c.alignment = CENTER
            hdr_c.border = BORDER
            
            for ci, mn in enumerate(finbit_months, 2):
                c = ws6.cell(hdr_row, ci, mn)
                c.font = HEADER_FONT_WHITE
                c.fill = HEADER_FILL
                c.alignment = CENTER
                c.border = BORDER
            
            # Metric rows: (dict_key, display_label, is_currency)
            FINBIT_ROWS = [
                ("monthlyAvgBal",     "Monthly Avg Balance",   True),
                ("maxBalance",        "Max Balance",           True),
                ("minBalance",        "Min Balance",           True),
                ("cashDeposit",       "Cash Deposits",         True),
                ("cashWithdrawals",   "Cash Withdrawals",      True),
                ("chqDeposit",        "Cheque Deposits",       True),
                ("chqIssues",         "Cheques Issued",        True),
                ("credits",           "Total Credits",         True),
                ("debits",            "Total Debits",          True),
                ("inwBounce",         "Inward Bounce",         False),
                ("outwBounce",        "Outward Bounce",        False),
                ("penaltyCharges",    "Penalty Charges",       True),
                ("ecsNach",           "ECS / NACH",            True),
                ("totalNetDebit",     "Total Net Debit",       True),
                ("totalNetCredit",    "Total Net Credit",      True),
                ("selfWithdraw",      "Self Withdrawal",       True),
                ("selfDeposit",       "Self Deposit",          True),
                ("loanRepayment",     "Loan Repayment",        True),
                ("loanCredit",        "Loan Credit",           True),
                ("creditCardPayment", "Credit Card Payment",   True),
                ("minCredits",        "Min Credit Amount",     True),
                ("maxCredits",        "Max Credit Amount",     True),
                ("salary",            "Salary",                True),
                ("bankCharges",       "Bank Charges",          True),
                None,  # section separator
                ("balanceOpening",    "BALANCE (Opening)",     True),
                ("balanceClosing",    "BALANCE (Closing)",     True),
                ("salaryMonth",       "SALARY (Income/Month)", True),
                ("ccPayment",         "CCPAYMENT",             True),
                ("eodMinBalance",     "EOD MIN BALANCE",       True),
                ("eodMaxBalance",     "EOD MAX BALANCE",       True),
            ]
            
            ALT_FILL = PatternFill(
                start_color="F2F7FB", end_color="F2F7FB",
                fill_type="solid")
            SEC_FILL = PatternFill(
                start_color="D6E4F0", end_color="D6E4F0",
                fill_type="solid")
            
            r = hdr_row + 1
            row_idx = 0
            for entry in FINBIT_ROWS:
                if entry is None:
                    # Derived section header
                    ws6.cell(r, 1, "Derived Monthly Metrics").font = Font(
                        bold=True, size=10, color="1F4E79")
                    ws6.cell(r, 1).fill = SEC_FILL
                    ws6.cell(r, 1).border = BORDER
                    for ci in range(2, len(finbit_months) + 2):
                        ws6.cell(r, ci).fill = SEC_FILL
                        ws6.cell(r, ci).border = BORDER
                    r += 1
                    row_idx = 0
                    continue
                
                key, label, is_cur = entry
                fill = ALT_FILL if row_idx % 2 == 0 else None
                
                lcell = ws6.cell(r, 1, label)
                lcell.font = Font(bold=True, size=10)
                lcell.border = BORDER
                lcell.alignment = LEFT
                if fill:
                    lcell.fill = fill
                
                for ci, mn in enumerate(finbit_months, 2):
                    val = finbit_data[mn].get(key, 0)
                    vc = ws6.cell(r, ci, val)
                    vc.border = BORDER
                    vc.alignment = RIGHT
                    if is_cur:
                        vc.number_format = CURRENCY
                    if fill:
                        vc.fill = fill
                
                r += 1
                row_idx += 1
            
            # Column widths
            ws6.column_dimensions['A'].width = 28
            for ci in range(2, len(finbit_months) + 2):
                ws6.column_dimensions[get_column_letter(ci)].width = 18
            
            ws6.freeze_panes = 'B5'
        
        # Save
        wb.save(output_path)
        self.logger.info("Excel saved: %s", output_path)
        
        return output_path
    
    def _aggregate_monthly(self, transactions: List[Dict]) -> Dict[str, Dict]:
        """
        Aggregate transactions by month with accurate opening balance calculation.
        
        Logic:
        - First month: Opening = First closing - First credit + First debit
        - Other months: Opening = Previous month's closing
        - Validation: Opening + Credits - Debits ≈ Closing (±1 rupee tolerance)
        """
        from datetime import datetime
        from collections import OrderedDict
        
        # Group transactions by month
        monthly_txns = defaultdict(list)
        
        for txn in transactions:
            date_str = txn.get('date', '')
            if not date_str:
                continue
            
            try:
                # Parse date
                if '-' in date_str:
                    dt = datetime.strptime(date_str, "%Y-%m-%d")
                else:
                    dt = datetime.strptime(date_str, "%d/%m/%Y")
                
                month_key = dt.strftime("%b-%y")
                monthly_txns[month_key].append({
                    'date': dt,
                    'credit': txn.get('credit') or 0,
                    'debit': txn.get('debit') or 0,
                    'balance': txn.get('balance', 0),
                })
            except:
                continue
        
        # Sort months chronologically
        sorted_months = sorted(monthly_txns.keys(), 
                              key=lambda m: datetime.strptime(m, "%b-%y"))
        
        result = OrderedDict()
        prev_closing = None
        
        for month_idx, month in enumerate(sorted_months):
            txns = sorted(monthly_txns[month], key=lambda t: t['date'])
            
            if not txns:
                continue
            
            # Calculate aggregates
            cr_count = sum(1 for t in txns if t['credit'] > 0)
            cr_amt = sum(t['credit'] for t in txns)
            dr_count = sum(1 for t in txns if t['debit'] > 0)
            dr_amt = sum(t['debit'] for t in txns)
            
            # Balance statistics
            balances = [t['balance'] for t in txns]
            avg_bal = sum(balances) / len(balances) if balances else 0
            min_bal = min(balances) if balances else 0
            max_bal = max(balances) if balances else 0
            end_bal = balances[-1] if balances else 0
            
            # Opening balance calculation
            if month_idx == 0:
                # First month: reconstruct from first transaction
                first_txn = txns[0]
                start_bal = first_txn['balance'] - first_txn['credit'] + first_txn['debit']
            else:
                # Subsequent months: use previous month's closing
                start_bal = prev_closing if prev_closing is not None else 0
            
            # Validation: Start + Credits - Debits ≈ End
            expected_end = start_bal + cr_amt - dr_amt
            reconciliation_diff = abs(expected_end - end_bal)
            
            if reconciliation_diff > 1.0:
                self.logger.warning(
                    "Month %s reconciliation mismatch: expected %.2f, got %.2f (diff: %.2f)",
                    month, expected_end, end_bal, reconciliation_diff
                )
            
            result[month] = {
                'cr_count': cr_count,
                'cr_amt': round(cr_amt, 2),
                'dr_count': dr_count,
                'dr_amt': round(dr_amt, 2),
                'avg_bal': round(avg_bal, 2),
                'min_bal': round(min_bal, 2),
                'max_bal': round(max_bal, 2),
                'start_bal': round(start_bal, 2),
                'end_bal': round(end_bal, 2),
                'reconciliation_diff': round(reconciliation_diff, 2),
            }
            
            prev_closing = end_bal
        
        return result
    
    def _aggregate_credit_categories(self, transactions: List[Dict]) -> Dict[str, float]:
        """Aggregate credit categories."""
        cats = {
            "UPI": 0.0,
            "Loan": 0.0,
            "Salary Credits": 0.0,
            "Bank Transfer": 0.0,
            "Cash Deposit": 0.0,
            "Others": 0.0,
        }
        
        for txn in transactions:
            credit = txn.get('credit')
            if not credit:
                continue
            
            cat = txn.get('category', 'Others')
            if cat in cats:
                cats[cat] += credit
            else:
                cats['Others'] += credit
        
        return {k: round(v, 2) for k, v in cats.items()}
    
    def _aggregate_debit_categories(self, transactions: List[Dict]) -> Dict[str, float]:
        """Aggregate debit categories."""
        cats = {
            "Loan Payments": 0.0,
            "ATM Withdrawal": 0.0,
            "Shopping": 0.0,
            "Bill Payment": 0.0,
            "Withdrawal": 0.0,
            "Investments": 0.0,
            "Others": 0.0,
        }
        
        for txn in transactions:
            debit = txn.get('debit')
            if not debit:
                continue
            
            cat = txn.get('category', 'Others')
            if cat in cats:
                cats[cat] += debit
            elif 'Others' in cat:
                cats['Others'] += debit
            else:
                cats['Others'] += debit
        
        return {k: round(v, 2) for k, v in cats.items()}
    
    def _aggregate_weekly_credit(self, transactions: List[Dict]) -> Dict[str, float]:
        """Aggregate credits by week of month."""
        weeks = {
            "Week1 (1-7)": 0.0,
            "Week2 (8-14)": 0.0,
            "Week3 (15-21)": 0.0,
            "Week4 (22-end)": 0.0,
        }
        
        for txn in transactions:
            credit = txn.get('credit')
            if not credit:
                continue
            
            date_str = txn.get('date', '')
            if not date_str:
                continue
            
            try:
                from datetime import datetime
                if '-' in date_str:
                    dt = datetime.strptime(date_str, "%Y-%m-%d")
                else:
                    dt = datetime.strptime(date_str, "%d/%m/%Y")
                
                day = dt.day
                if 1 <= day <= 7:
                    weeks["Week1 (1-7)"] += credit
                elif 8 <= day <= 14:
                    weeks["Week2 (8-14)"] += credit
                elif 15 <= day <= 21:
                    weeks["Week3 (15-21)"] += credit
                else:
                    weeks["Week4 (22-end)"] += credit
            except:
                continue
        
        return {k: round(v, 2) for k, v in weeks.items()}
    
    def _aggregate_weekly_debit(self, transactions: List[Dict]) -> Dict[str, float]:
        """Aggregate debits by week of month."""
        weeks = {
            "Week1 (1-7)": 0.0,
            "Week2 (8-14)": 0.0,
            "Week3 (15-21)": 0.0,
            "Week4 (22-end)": 0.0,
        }
        
        for txn in transactions:
            debit = txn.get('debit')
            if not debit:
                continue
            
            date_str = txn.get('date', '')
            if not date_str:
                continue
            
            try:
                from datetime import datetime
                if '-' in date_str:
                    dt = datetime.strptime(date_str, "%Y-%m-%d")
                else:
                    dt = datetime.strptime(date_str, "%d/%m/%Y")
                
                day = dt.day
                if 1 <= day <= 7:
                    weeks["Week1 (1-7)"] += debit
                elif 8 <= day <= 14:
                    weeks["Week2 (8-14)"] += debit
                elif 15 <= day <= 21:
                    weeks["Week3 (15-21)"] += debit
                else:
                    weeks["Week4 (22-end)"] += debit
            except:
                continue
        
        return {k: round(v, 2) for k, v in weeks.items()}
    
    def _aggregate_recurring(self, transactions: List[Dict]) -> Dict[str, float]:
        """Aggregate recurring vs non-recurring."""
        stats = {
            'cr_recurring': 0.0,
            'cr_non_recurring': 0.0,
            'cr_total': 0.0,
            'dr_recurring': 0.0,
            'dr_non_recurring': 0.0,
            'dr_total': 0.0,
        }
        
        for txn in transactions:
            is_recurring = txn.get('is_recurring', False)
            credit = txn.get('credit') or 0
            debit = txn.get('debit') or 0
            
            if credit > 0:
                stats['cr_total'] += credit
                if is_recurring:
                    stats['cr_recurring'] += credit
                else:
                    stats['cr_non_recurring'] += credit
            
            if debit > 0:
                stats['dr_total'] += debit
                if is_recurring:
                    stats['dr_recurring'] += debit
                else:
                    stats['dr_non_recurring'] += debit
        
        return {k: round(v, 2) for k, v in stats.items()}
    
    def _compute_finbit_monthly(
        self, transactions: List[Dict], opening_balance: float = 0
    ) -> tuple:
        """
        Compute all Finbit monthly metrics from transactions.
        
        Returns:
            (sorted_month_keys, OrderedDict of month -> metrics_dict)
        """
        from datetime import datetime
        from collections import OrderedDict
        
        def _kw(desc: str, keywords: list) -> bool:
            """Case-insensitive keyword match."""
            d = desc.upper()
            return any(k.upper() in d for k in keywords)
        
        # ---- keyword lists for transaction-type detection ----
        KW_CASH_DEP    = ["CASH DEPOSIT", "CASHDEP", "CDM", "CASHDEPOSITBY",
                          "CASH DEP", "BY CASH"]
        KW_CASH_WDL    = ["ATW", "ATM WDL", "ATM CASH", "ATMWDL",
                          "CASH WITHDRAWAL", "CASHWITHDRAWAL", "NFS ATM",
                          "ATM-WDL", "ATM/CASH", "CASH W/D"]
        KW_CHQ_DEP     = ["CHQ DEP", "CHEQUE DEP", "CLG CR", "I/WCLG",
                          "INWARD CLG", "IW CLR", "CHQDEP"]
        KW_CHQ_ISS     = ["CHQPAID", "CHQ PAID", "SELF-CHQ", "CLG DR",
                          "O/WCLG", "OUTWARD CLG", "CHEQUE PAID"]
        KW_INW_BOUNCE  = ["I/WCHQRET", "INWARD RETURN", "INW BOUNCE",
                          "INW RET", "INWARD BOUNCE"]
        KW_OUTW_BOUNCE = ["O/WCHQRET", "OUTWARD RETURN", "OUTW BOUNCE",
                          "O/W RETURN", "OUTWARD BOUNCE"]
        KW_PENALTY     = ["PENALTY", "PENAL CHARGE", "PENAL INT",
                          "MIN BAL CHARGE", "NON-MAINT", "MINIMUM BALANCE"]
        KW_ECS_NACH    = ["ACHD-", "ACH D-", "ECS/", "ECS ", "NACH/",
                          "NACH ", "AUTOPAYSI", "SI-", "AUTO DEBIT", "SI /"]
        KW_SELF        = ["SELF-", "/SELF", " SELF ", "SELF CHQ", "SELF TRF"]
        KW_LOAN_REP    = ["EMI", "LOAN REPAY", "LOAN EMI"]
        KW_LOAN_CR     = ["LOAN DISBURSE", "LOAN CR", "LOAN SANCTION"]
        KW_CC_PAY      = ["CC PAYMENT", "CREDIT CARD", "CC000",
                          "RAZPCREDCLUB", "CRED CLUB", "CREDITCARD", "CCPAY"]
        KW_SALARY      = ["SALARY", "SAL CR", "PAYROLL", "WAGES", "STIPEND"]
        KW_BANK_CHG    = ["CHARGES", "FEE-", "LOWUSAGECHARGES",
                          "SETTLEMENTCHARGE", "EDC RENTAL", "EDCRENTAL",
                          "SERVICE CHARGE", "BANK CHARGE", "SMS ALERT",
                          "MAINTENANCE CHARGE", "FEE-ATMCASH"]
        
        # Group transactions by month
        monthly_txns = defaultdict(list)
        for txn in transactions:
            date_str = txn.get('date', '')
            if not date_str:
                continue
            try:
                if '-' in date_str:
                    dt = datetime.strptime(date_str, "%Y-%m-%d")
                else:
                    dt = datetime.strptime(date_str, "%d/%m/%Y")
                monthly_txns[dt.strftime("%b-%y")].append({**txn, '_dt': dt})
            except Exception:
                continue
        
        sorted_months = sorted(
            monthly_txns.keys(),
            key=lambda m: datetime.strptime(m, "%b-%y"),
        )
        
        result = OrderedDict()
        prev_closing = opening_balance
        
        for month_idx, month in enumerate(sorted_months):
            txns = sorted(monthly_txns[month], key=lambda t: t['_dt'])
            if not txns:
                continue
            
            # ---- basic aggregates ----
            cr_vals = [(t.get('credit') or 0) for t in txns
                       if (t.get('credit') or 0) > 0]
            dr_vals = [(t.get('debit') or 0) for t in txns
                       if (t.get('debit') or 0) > 0]
            total_cr = sum(cr_vals)
            total_dr = sum(dr_vals)
            
            # ---- balance analytics ----
            balances = [t.get('balance', 0) for t in txns]
            
            # EOD: last closing balance per calendar day
            eod = {}
            for t in txns:
                eod[t['_dt'].date()] = t.get('balance', 0)
            eod_vals = list(eod.values()) if eod else [0]
            
            # Opening / Closing
            if month_idx == 0:
                f = txns[0]
                start_bal = (f.get('balance', 0)
                             - (f.get('credit') or 0)
                             + (f.get('debit') or 0))
            else:
                start_bal = prev_closing
            end_bal = balances[-1] if balances else prev_closing
            
            # ---- transaction type detection ----
            desc = lambda t: t.get('description', '')
            cat  = lambda t: t.get('category', '')
            cr   = lambda t: (t.get('credit') or 0)
            dr   = lambda t: (t.get('debit') or 0)
            
            cash_dep = sum(cr(t) for t in txns
                          if cr(t) > 0 and (cat(t) == 'Cash Deposit'
                             or _kw(desc(t), KW_CASH_DEP)))
            cash_wdl = sum(dr(t) for t in txns
                          if dr(t) > 0 and (cat(t) == 'ATM Withdrawal'
                             or _kw(desc(t), KW_CASH_WDL)))
            chq_dep  = sum(cr(t) for t in txns
                          if cr(t) > 0 and _kw(desc(t), KW_CHQ_DEP))
            chq_iss  = sum(dr(t) for t in txns
                          if dr(t) > 0 and _kw(desc(t), KW_CHQ_ISS))
            inw_b    = sum(1 for t in txns
                          if _kw(desc(t), KW_INW_BOUNCE))
            outw_b   = sum(1 for t in txns
                          if _kw(desc(t), KW_OUTW_BOUNCE))
            penalty  = sum(dr(t) for t in txns
                          if dr(t) > 0 and _kw(desc(t), KW_PENALTY))
            ecs_nach = sum(dr(t) for t in txns
                          if dr(t) > 0 and _kw(desc(t), KW_ECS_NACH))
            self_wdl = sum(dr(t) for t in txns
                          if dr(t) > 0 and _kw(desc(t), KW_SELF))
            self_dep = sum(cr(t) for t in txns
                          if cr(t) > 0 and _kw(desc(t), KW_SELF))
            loan_rep = sum(dr(t) for t in txns
                          if dr(t) > 0 and (cat(t) == 'Loan Payments'
                             or _kw(desc(t), KW_LOAN_REP)))
            loan_crd = sum(cr(t) for t in txns
                          if cr(t) > 0 and (cat(t) == 'Loan'
                             or _kw(desc(t), KW_LOAN_CR)))
            cc_pay   = sum(dr(t) for t in txns
                          if dr(t) > 0 and _kw(desc(t), KW_CC_PAY))
            sal      = sum(cr(t) for t in txns
                          if cr(t) > 0 and (cat(t) == 'Salary Credits'
                             or _kw(desc(t), KW_SALARY)))
            bank_chg = sum(dr(t) for t in txns
                          if dr(t) > 0 and _kw(desc(t), KW_BANK_CHG))
            
            result[month] = {
                'monthlyAvgBal':      round(sum(eod_vals) / len(eod_vals), 2),
                'maxBalance':         round(max(balances), 2),
                'minBalance':         round(min(balances), 2),
                'cashDeposit':        round(cash_dep, 2),
                'cashWithdrawals':    round(cash_wdl, 2),
                'chqDeposit':         round(chq_dep, 2),
                'chqIssues':          round(chq_iss, 2),
                'credits':            round(total_cr, 2),
                'debits':             round(total_dr, 2),
                'inwBounce':          inw_b,
                'outwBounce':         outw_b,
                'penaltyCharges':     round(penalty, 2),
                'ecsNach':            round(ecs_nach, 2),
                'totalNetDebit':      round(max(total_dr - total_cr, 0), 2),
                'totalNetCredit':     round(max(total_cr - total_dr, 0), 2),
                'selfWithdraw':       round(self_wdl, 2),
                'selfDeposit':        round(self_dep, 2),
                'loanRepayment':      round(loan_rep, 2),
                'loanCredit':         round(loan_crd, 2),
                'creditCardPayment':  round(cc_pay, 2),
                'minCredits':         round(min(cr_vals), 2) if cr_vals else 0,
                'maxCredits':         round(max(cr_vals), 2) if cr_vals else 0,
                'salary':             round(sal, 2),
                'bankCharges':        round(bank_chg, 2),
                'balanceOpening':     round(start_bal, 2),
                'balanceClosing':     round(end_bal, 2),
                'salaryMonth':        round(sal, 2),
                'ccPayment':          round(cc_pay, 2),
                'eodMinBalance':      round(min(eod_vals), 2),
                'eodMaxBalance':      round(max(eod_vals), 2),
            }
            prev_closing = end_bal
        
        return sorted_months, result

"""
Formula-Based Excel Report Engine
==================================
Generates Excel reports using ONLY Excel formulas for all calculations.
No hardcoded numeric values - everything is formula-driven.

Column Mapping (FIXED - DO NOT CHANGE):
  Column A = Date         (date values)
  Column B = Description  (text)
  Column C = Debit        (numeric — blank if no debit)
  Column D = Credit       (numeric — blank if no credit)
  Column E = Balance      (numeric)
  Column F = Category     (text)
  Column G = Confidence   (text or %)
  Column H = Recurring    (text: "Yes" or "No")

ABSOLUTE RULES:
  - Column D (Credit) = ONLY source for ALL credit calculations
  - Column C (Debit) = ONLY source for ALL debit calculations
  - NEVER infer credit/debit from Description or Balance
  - Blank cell = zero. Never skip. Never guess.
"""

import logging
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
from calendar import monthrange
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class FormulaExcelEngine:
    """
    Excel report generator using 100% Excel formulas.
    No Python-calculated values in output sheets.
    """
    
    # Fixed column mapping
    COL_DATE = 'A'
    COL_DESC = 'B'
    COL_DEBIT = 'C'
    COL_CREDIT = 'D'
    COL_BALANCE = 'E'
    COL_CATEGORY = 'F'
    COL_CONFIDENCE = 'G'
    COL_RECURRING = 'H'
    
    # Sheet name
    RAW_SHEET = 'Raw Transactions'
    
    # Styles
    FONT_DEFAULT = Font(name='Arial', size=10)
    FONT_BOLD = Font(name='Arial', size=10, bold=True)
    FONT_HEADER = Font(name='Arial', size=10, bold=True)
    
    ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
    ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
    ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
    
    BORDER_THIN = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    FILL_LIGHT_BLUE = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    FILL_LIGHT_ORANGE = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    
    # Number formats
    FMT_CURRENCY = '₹#,##0.00'
    FMT_INTEGER = '#,##0'
    FMT_DATE = 'DD-MM-YYYY'
    
    def __init__(self):
        self.workbook = None
        self.last_row = 0
        self.months: List[Tuple[int, int]] = []
    
    def generate(
        self,
        transactions: List[Dict[str, Any]],
        metadata: Dict[str, Any] = None,
        output_path: str = None
    ) -> bytes:
        """
        Generate Excel report with formula-based calculations.
        
        Args:
            transactions: List of transaction dicts with keys:
                date, description, debit, credit, balance, category, confidence, recurring
            metadata: Optional dict with name, account_no
            output_path: Optional file path to save
            
        Returns:
            Excel file bytes
        """
        self.workbook = Workbook()
        metadata = metadata or {}
        
        # Calculate last_row (header + data rows)
        self.last_row = len(transactions) + 1
        
        # Extract unique months from transactions
        self._extract_months(transactions)
        
        # Create sheets in order
        # Remove default sheet first
        default_sheet = self.workbook.active
        
        # Create all sheets
        ws_summary = self.workbook.create_sheet("Summary", 0)
        ws_category = self.workbook.create_sheet("Category Analysis", 1)
        ws_weekly = self.workbook.create_sheet("Weekly Analysis", 2)
        ws_recurring = self.workbook.create_sheet("Recurring Analysis", 3)
        ws_raw = self.workbook.create_sheet(self.RAW_SHEET, 4)
        
        # Remove default sheet
        self.workbook.remove(default_sheet)
        
        # Build sheets
        self._build_raw_transactions(ws_raw, transactions)
        self._build_summary(ws_summary, metadata)
        self._build_category_analysis(ws_category)
        self._build_weekly_analysis(ws_weekly)
        self._build_recurring_analysis(ws_recurring)
        
        # Save to bytes
        from io import BytesIO
        buffer = BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        
        if output_path:
            with open(output_path, 'wb') as f:
                f.write(buffer.getvalue())
        
        return buffer.getvalue()
    
    def _extract_months(self, transactions: List[Dict[str, Any]]):
        """Extract unique (year, month) tuples from transactions, sorted chronologically."""
        months_set = set()
        
        for txn in transactions:
            date_val = txn.get('date')
            if date_val:
                try:
                    if isinstance(date_val, str):
                        # Try multiple date formats
                        for fmt in ['%d/%m/%y', '%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                            try:
                                dt = datetime.strptime(date_val, fmt)
                                months_set.add((dt.year, dt.month))
                                break
                            except ValueError:
                                continue
                    elif isinstance(date_val, datetime):
                        months_set.add((date_val.year, date_val.month))
                except Exception:
                    pass
        
        self.months = sorted(list(months_set))
    
    def _get_last_day(self, year: int, month: int) -> int:
        """Get the last day of a month."""
        return monthrange(year, month)[1]
    
    def _build_raw_transactions(self, ws, transactions: List[Dict[str, Any]]):
        """Build the Raw Transactions sheet with source data."""
        # Headers
        headers = ['Date', 'Description', 'Debit', 'Credit', 'Balance', 'Category', 'Confidence', 'Recurring']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.FONT_BOLD
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
        
        # Data rows
        for row_idx, txn in enumerate(transactions, 2):
            # Column A - Date
            date_val = txn.get('date', '')
            if isinstance(date_val, str):
                # Try to parse and convert to datetime
                for fmt in ['%d/%m/%y', '%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                    try:
                        date_val = datetime.strptime(date_val, fmt)
                        break
                    except ValueError:
                        continue
            ws.cell(row=row_idx, column=1, value=date_val).number_format = self.FMT_DATE
            
            # Column B - Description
            ws.cell(row=row_idx, column=2, value=txn.get('description', ''))
            
            # Column C - Debit
            debit = txn.get('debit')
            if debit is not None and debit > 0:
                ws.cell(row=row_idx, column=3, value=debit).number_format = self.FMT_CURRENCY
            
            # Column D - Credit
            credit = txn.get('credit')
            if credit is not None and credit > 0:
                ws.cell(row=row_idx, column=4, value=credit).number_format = self.FMT_CURRENCY
            
            # Column E - Balance
            balance = txn.get('balance')
            if balance is not None:
                ws.cell(row=row_idx, column=5, value=balance).number_format = self.FMT_CURRENCY
            
            # Column F - Category
            ws.cell(row=row_idx, column=6, value=txn.get('category', ''))
            
            # Column G - Confidence
            ws.cell(row=row_idx, column=7, value=txn.get('confidence', ''))
            
            # Column H - Recurring
            ws.cell(row=row_idx, column=8, value=txn.get('recurring', 'No'))
        
        # Apply formatting
        for row in ws.iter_rows(min_row=2, max_row=self.last_row, min_col=1, max_col=8):
            for cell in row:
                cell.font = self.FONT_DEFAULT
                cell.border = self.BORDER_THIN
        
        # Column widths
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 13
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _build_summary(self, ws, metadata: Dict[str, Any]):
        """Build the Summary sheet with header block and monthly table."""
        lr = self.last_row
        RAW = f"'{self.RAW_SHEET}'"
        
        # ═══════════════════════════════════════════════════
        # HEADER BLOCK (Rows 1-4)
        # ═══════════════════════════════════════════════════
        
        header_data = [
            ('Name', metadata.get('name', '')),
            ('Account No', metadata.get('account_no', '')),
            ('Statement From', f"=MIN({RAW}!A:A)"),
            ('Statement To', f"=MAX({RAW}!A:A)")
        ]
        
        for row_idx, (label, value) in enumerate(header_data, 1):
            # Label
            cell_label = ws.cell(row=row_idx, column=1, value=label)
            cell_label.font = self.FONT_BOLD
            cell_label.alignment = self.ALIGN_LEFT
            cell_label.border = self.BORDER_THIN
            
            # Value
            cell_value = ws.cell(row=row_idx, column=2, value=value)
            cell_value.font = self.FONT_DEFAULT
            cell_value.alignment = self.ALIGN_LEFT
            cell_value.border = self.BORDER_THIN
            
            # Date format for rows 3-4
            if row_idx >= 3:
                cell_value.number_format = self.FMT_DATE
        
        # ═══════════════════════════════════════════════════
        # MONTHLY TABLE (Row 6 onward)
        # ═══════════════════════════════════════════════════
        
        # Row labels
        row_labels = [
            '',  # Row 6 - Header row (blank in column A)
            'Total Credit Count',
            'Total Credit Amount',
            'Total Debit Count',
            'Total Debit Amount',
            'Avg Balance',
            'Min Balance',
            'Max Balance',
            'Start of Month Balance',
            'End of Month Balance'
        ]
        
        # Write row labels
        for row_offset, label in enumerate(row_labels):
            row_idx = 6 + row_offset
            cell = ws.cell(row=row_idx, column=1, value=label)
            cell.font = self.FONT_BOLD if row_offset == 0 else self.FONT_BOLD
            cell.alignment = self.ALIGN_LEFT
            cell.border = self.BORDER_THIN
        
        # Write month columns
        for col_offset, (year, month) in enumerate(self.months):
            col_idx = 2 + col_offset
            col_letter = get_column_letter(col_idx)
            
            # Month label (row 6)
            month_name = datetime(year, month, 1).strftime('%b %Y')
            cell_header = ws.cell(row=6, column=col_idx, value=month_name)
            cell_header.font = self.FONT_BOLD
            cell_header.alignment = self.ALIGN_CENTER
            cell_header.border = self.BORDER_THIN
            cell_header.fill = self.FILL_LIGHT_BLUE
            
            # Calculate last day of month
            last_day = self._get_last_day(year, month)
            
            # Date range conditions for formulas
            date_gte = f'">="&DATE({year},{month},1)'
            date_lte = f'"<="&DATE({year},{month},{last_day})'
            
            # Row 7 - Total Credit Count
            formula = f'=COUNTIFS({RAW}!A$2:A${lr},{date_gte},{RAW}!A$2:A${lr},{date_lte},{RAW}!D$2:D${lr},">"&0)'
            cell = ws.cell(row=7, column=col_idx, value=formula)
            cell.number_format = self.FMT_INTEGER
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Row 8 - Total Credit Amount
            formula = f'=SUMIFS({RAW}!D$2:D${lr},{RAW}!A$2:A${lr},{date_gte},{RAW}!A$2:A${lr},{date_lte})'
            cell = ws.cell(row=8, column=col_idx, value=formula)
            cell.number_format = self.FMT_CURRENCY
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Row 9 - Total Debit Count
            formula = f'=COUNTIFS({RAW}!A$2:A${lr},{date_gte},{RAW}!A$2:A${lr},{date_lte},{RAW}!C$2:C${lr},">"&0)'
            cell = ws.cell(row=9, column=col_idx, value=formula)
            cell.number_format = self.FMT_INTEGER
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Row 10 - Total Debit Amount
            formula = f'=SUMIFS({RAW}!C$2:C${lr},{RAW}!A$2:A${lr},{date_gte},{RAW}!A$2:A${lr},{date_lte})'
            cell = ws.cell(row=10, column=col_idx, value=formula)
            cell.number_format = self.FMT_CURRENCY
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Row 11 - Avg Balance
            formula = f'=AVERAGEIFS({RAW}!E$2:E${lr},{RAW}!A$2:A${lr},{date_gte},{RAW}!A$2:A${lr},{date_lte})'
            cell = ws.cell(row=11, column=col_idx, value=formula)
            cell.number_format = self.FMT_CURRENCY
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Row 12 - Min Balance
            formula = f'=MINIFS({RAW}!E$2:E${lr},{RAW}!A$2:A${lr},{date_gte},{RAW}!A$2:A${lr},{date_lte})'
            cell = ws.cell(row=12, column=col_idx, value=formula)
            cell.number_format = self.FMT_CURRENCY
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Row 13 - Max Balance
            formula = f'=MAXIFS({RAW}!E$2:E${lr},{RAW}!A$2:A${lr},{date_gte},{RAW}!A$2:A${lr},{date_lte})'
            cell = ws.cell(row=13, column=col_idx, value=formula)
            cell.number_format = self.FMT_CURRENCY
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Row 14 - Start of Month Balance (first balance in that month)
            # Use simpler MINIFS approach to get balance on first date of month
            formula = (
                f"=IFERROR(INDEX({RAW}!E$2:E${lr},"
                f"MATCH(MINIFS({RAW}!A$2:A${lr},{RAW}!A$2:A${lr},\">=\"&DATE({year},{month},1),"
                f"{RAW}!A$2:A${lr},\"<=\"&DATE({year},{month},{last_day})),"
                f"{RAW}!A$2:A${lr},0)),\"\")"
            )
            cell = ws.cell(row=14, column=col_idx, value=formula)
            cell.number_format = self.FMT_CURRENCY
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Row 15 - End of Month Balance (last balance in that month)
            # Use MAXIFS to find last date, then INDEX/MATCH to get balance
            formula = (
                f"=IFERROR(INDEX({RAW}!E$2:E${lr},"
                f"MATCH(MAXIFS({RAW}!A$2:A${lr},{RAW}!A$2:A${lr},\">=\"&DATE({year},{month},1),"
                f"{RAW}!A$2:A${lr},\"<=\"&DATE({year},{month},{last_day})),"
                f"{RAW}!A$2:A${lr},0)),\"\")"
            )
            cell = ws.cell(row=15, column=col_idx, value=formula)
            cell.number_format = self.FMT_CURRENCY
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
        
        # Column widths
        ws.column_dimensions['A'].width = 28
        for col_offset in range(len(self.months)):
            col_letter = get_column_letter(2 + col_offset)
            ws.column_dimensions[col_letter].width = 18
        
        # Freeze panes
        ws.freeze_panes = 'A7'
    
    def _build_category_analysis(self, ws):
        """Build the Category Analysis sheet with credit and debit category tables."""
        lr = self.last_row
        RAW = f"'{self.RAW_SHEET}'"
        
        # ═══════════════════════════════════════════════════
        # LEFT TABLE - Credit Categories (Column A-C)
        # ═══════════════════════════════════════════════════
        
        # Header row
        headers_left = ['Credit Categories', 'Amount', 'Count']
        for col, header in enumerate(headers_left, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.FONT_BOLD
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            cell.fill = self.FILL_LIGHT_BLUE
        
        # Credit categories
        credit_categories = [
            'UPI',
            'Loan',
            'Salary Credits',
            'Bank Transfer',
            'Cash Deposit',
            'Others Credit',
            'Total Credit Amount'
        ]
        
        for row_offset, category in enumerate(credit_categories):
            row_idx = 2 + row_offset
            
            # Category name
            cell = ws.cell(row=row_idx, column=1, value=category)
            cell.font = self.FONT_BOLD if 'Total' in category else self.FONT_DEFAULT
            cell.alignment = self.ALIGN_LEFT
            cell.border = self.BORDER_THIN
            
            if 'Total' in category:
                # Total row - sum of above
                amount_formula = f'=SUM(B2:B{row_idx-1})'
                count_formula = f'=SUM(C2:C{row_idx-1})'
            else:
                # Category row
                amount_formula = f'=SUMIFS({RAW}!D$2:D${lr},{RAW}!F$2:F${lr},"*{category}*")'
                count_formula = f'=COUNTIFS({RAW}!F$2:F${lr},"*{category}*",{RAW}!D$2:D${lr},">"&0)'
            
            # Amount
            cell = ws.cell(row=row_idx, column=2, value=amount_formula)
            cell.number_format = self.FMT_CURRENCY
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Count
            cell = ws.cell(row=row_idx, column=3, value=count_formula)
            cell.number_format = self.FMT_INTEGER
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
        
        # ═══════════════════════════════════════════════════
        # RIGHT TABLE - Debit Categories (Column F-H)
        # ═══════════════════════════════════════════════════
        
        # Header row
        headers_right = ['Debit Categories', 'Amount', 'Count']
        for col_offset, header in enumerate(headers_right):
            col = 6 + col_offset  # F, G, H
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.FONT_BOLD
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            cell.fill = self.FILL_LIGHT_ORANGE
        
        # Debit categories
        debit_categories = [
            'Loan Payments',
            'ATM Withdrawal',
            'Shopping',
            'Bill Payment',
            'Withdrawal',
            'Investments',
            'Others Debit',
            'Total Debit Amount'
        ]
        
        for row_offset, category in enumerate(debit_categories):
            row_idx = 2 + row_offset
            
            # Category name
            cell = ws.cell(row=row_idx, column=6, value=category)
            cell.font = self.FONT_BOLD if 'Total' in category else self.FONT_DEFAULT
            cell.alignment = self.ALIGN_LEFT
            cell.border = self.BORDER_THIN
            
            if 'Total' in category:
                # Total row - sum of above
                amount_formula = f'=SUM(G2:G{row_idx-1})'
                count_formula = f'=SUM(H2:H{row_idx-1})'
            else:
                # Category row
                amount_formula = f'=SUMIFS({RAW}!C$2:C${lr},{RAW}!F$2:F${lr},"*{category}*")'
                count_formula = f'=COUNTIFS({RAW}!F$2:F${lr},"*{category}*",{RAW}!C$2:C${lr},">"&0)'
            
            # Amount
            cell = ws.cell(row=row_idx, column=7, value=amount_formula)
            cell.number_format = self.FMT_CURRENCY
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Count
            cell = ws.cell(row=row_idx, column=8, value=count_formula)
            cell.number_format = self.FMT_INTEGER
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
        
        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['E'].width = 3  # Spacer
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 10
    
    def _build_weekly_analysis(self, ws):
        """Build the Weekly Analysis sheet with SUMPRODUCT+DAY formulas."""
        lr = self.last_row
        RAW = f"'{self.RAW_SHEET}'"
        
        # ═══════════════════════════════════════════════════
        # LEFT TABLE - Credit Weekly Split (Column A-C)
        # ═══════════════════════════════════════════════════
        
        # Header row
        headers_left = ['Week', 'Credit Amount', 'Credit Count']
        for col, header in enumerate(headers_left, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.FONT_BOLD
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            cell.fill = self.FILL_LIGHT_BLUE
        
        # Week definitions
        weeks = [
            ('Week 1 (Days 1-7)', 1, 7),
            ('Week 2 (Days 8-14)', 8, 14),
            ('Week 3 (Days 15-21)', 15, 21),
            ('Week 4 (Days 22-31)', 22, 31),
            ('Total', None, None)
        ]
        
        for row_offset, (label, day_start, day_end) in enumerate(weeks):
            row_idx = 2 + row_offset
            
            # Week label
            cell = ws.cell(row=row_idx, column=1, value=label)
            cell.font = self.FONT_BOLD if label == 'Total' else self.FONT_DEFAULT
            cell.alignment = self.ALIGN_LEFT
            cell.border = self.BORDER_THIN
            
            if label == 'Total':
                # Sum of above
                amount_formula = f'=SUM(B2:B{row_idx-1})'
                count_formula = f'=SUM(C2:C{row_idx-1})'
            else:
                # SUMPRODUCT formula for week
                amount_formula = f'=SUMPRODUCT(({RAW}!D$2:D${lr})*(DAY({RAW}!A$2:A${lr})>={day_start})*(DAY({RAW}!A$2:A${lr})<={day_end}))'
                count_formula = f'=SUMPRODUCT(({RAW}!D$2:D${lr}>0)*(DAY({RAW}!A$2:A${lr})>={day_start})*(DAY({RAW}!A$2:A${lr})<={day_end})*1)'
            
            # Amount
            cell = ws.cell(row=row_idx, column=2, value=amount_formula)
            cell.number_format = self.FMT_CURRENCY
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Count
            cell = ws.cell(row=row_idx, column=3, value=count_formula)
            cell.number_format = self.FMT_INTEGER
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
        
        # ═══════════════════════════════════════════════════
        # RIGHT TABLE - Debit Weekly Split (Column F-H)
        # ═══════════════════════════════════════════════════
        
        # Header row
        headers_right = ['Week', 'Debit Amount', 'Debit Count']
        for col_offset, header in enumerate(headers_right):
            col = 6 + col_offset
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.FONT_BOLD
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            cell.fill = self.FILL_LIGHT_ORANGE
        
        for row_offset, (label, day_start, day_end) in enumerate(weeks):
            row_idx = 2 + row_offset
            
            # Week label
            cell = ws.cell(row=row_idx, column=6, value=label)
            cell.font = self.FONT_BOLD if label == 'Total' else self.FONT_DEFAULT
            cell.alignment = self.ALIGN_LEFT
            cell.border = self.BORDER_THIN
            
            if label == 'Total':
                # Sum of above
                amount_formula = f'=SUM(G2:G{row_idx-1})'
                count_formula = f'=SUM(H2:H{row_idx-1})'
            else:
                # SUMPRODUCT formula for week
                amount_formula = f'=SUMPRODUCT(({RAW}!C$2:C${lr})*(DAY({RAW}!A$2:A${lr})>={day_start})*(DAY({RAW}!A$2:A${lr})<={day_end}))'
                count_formula = f'=SUMPRODUCT(({RAW}!C$2:C${lr}>0)*(DAY({RAW}!A$2:A${lr})>={day_start})*(DAY({RAW}!A$2:A${lr})<={day_end})*1)'
            
            # Amount
            cell = ws.cell(row=row_idx, column=7, value=amount_formula)
            cell.number_format = self.FMT_CURRENCY
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Count
            cell = ws.cell(row=row_idx, column=8, value=count_formula)
            cell.number_format = self.FMT_INTEGER
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
        
        # Column widths
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['E'].width = 3  # Spacer
        ws.column_dimensions['F'].width = 22
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 14
    
    def _build_recurring_analysis(self, ws):
        """Build the Recurring Analysis sheet."""
        lr = self.last_row
        RAW = f"'{self.RAW_SHEET}'"
        
        # ═══════════════════════════════════════════════════
        # LEFT TABLE - Credit Recurring (Column A-C)
        # ═══════════════════════════════════════════════════
        
        # Header row
        headers_left = ['Type', 'Credit Amount', 'Credit Count']
        for col, header in enumerate(headers_left, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.FONT_BOLD
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            cell.fill = self.FILL_LIGHT_BLUE
        
        # Recurring types
        recurring_types = [
            ('Recurring', 'Yes'),
            ('Non-Recurring', 'No'),
            ('Total', None)
        ]
        
        for row_offset, (label, flag) in enumerate(recurring_types):
            row_idx = 2 + row_offset
            
            # Type label
            cell = ws.cell(row=row_idx, column=1, value=label)
            cell.font = self.FONT_BOLD if label == 'Total' else self.FONT_DEFAULT
            cell.alignment = self.ALIGN_LEFT
            cell.border = self.BORDER_THIN
            
            if label == 'Total':
                # Sum of above
                amount_formula = '=SUM(B2:B3)'
                count_formula = '=SUM(C2:C3)'
            else:
                # SUMIFS/COUNTIFS formula
                amount_formula = f'=SUMIFS({RAW}!D$2:D${lr},{RAW}!H$2:H${lr},"{flag}")'
                count_formula = f'=COUNTIFS({RAW}!H$2:H${lr},"{flag}",{RAW}!D$2:D${lr},">"&0)'
            
            # Amount
            cell = ws.cell(row=row_idx, column=2, value=amount_formula)
            cell.number_format = self.FMT_CURRENCY
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Count
            cell = ws.cell(row=row_idx, column=3, value=count_formula)
            cell.number_format = self.FMT_INTEGER
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
        
        # ═══════════════════════════════════════════════════
        # RIGHT TABLE - Debit Recurring (Column F-H)
        # ═══════════════════════════════════════════════════
        
        # Header row
        headers_right = ['Type', 'Debit Amount', 'Debit Count']
        for col_offset, header in enumerate(headers_right):
            col = 6 + col_offset
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.FONT_BOLD
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            cell.fill = self.FILL_LIGHT_ORANGE
        
        for row_offset, (label, flag) in enumerate(recurring_types):
            row_idx = 2 + row_offset
            
            # Type label
            cell = ws.cell(row=row_idx, column=6, value=label)
            cell.font = self.FONT_BOLD if label == 'Total' else self.FONT_DEFAULT
            cell.alignment = self.ALIGN_LEFT
            cell.border = self.BORDER_THIN
            
            if label == 'Total':
                # Sum of above
                amount_formula = '=SUM(G2:G3)'
                count_formula = '=SUM(H2:H3)'
            else:
                # SUMIFS/COUNTIFS formula
                amount_formula = f'=SUMIFS({RAW}!C$2:C${lr},{RAW}!H$2:H${lr},"{flag}")'
                count_formula = f'=COUNTIFS({RAW}!H$2:H${lr},"{flag}",{RAW}!C$2:C${lr},">"&0)'
            
            # Amount
            cell = ws.cell(row=row_idx, column=7, value=amount_formula)
            cell.number_format = self.FMT_CURRENCY
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Count
            cell = ws.cell(row=row_idx, column=8, value=count_formula)
            cell.number_format = self.FMT_INTEGER
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
        
        # Column widths
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['E'].width = 3  # Spacer
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 14
